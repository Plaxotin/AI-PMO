#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Экспорт Plan в .xlsx (BL-1): пользовательская конвертация .mpp → Excel.

Заголовки взяты из plan_parser.COLUMN_SYNONYMS, чтобы экспортированный
файл гарантированно читался обратно plan_parser'ом (round-trip).
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

HEADERS = [
    ('ID', 7), ('WBS', 11), ('Название', 60), ('Тип', 10),
    ('Начало', 12), ('Окончание', 12), ('Длительность', 13),
    ('% завершения', 14), ('Предшественники', 16), ('Крайний срок', 12),
    ('Затраты', 12), ('Базовое начало', 14), ('Базовое окончание', 16),
    ('Ответственный', 24),
]
DATE_COLS = (5, 6, 10, 12, 13)  # 1-based: Начало, Окончание, Крайний срок, Баз.*


def plan_to_xlsx(plan, path: str) -> str:
    """Сохраняет Plan в Excel и возвращает путь к файлу."""
    wb = Workbook()
    ws = wb.active
    ws.title = 'План'

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='44546A')
    for col, (title, width) in enumerate(HEADERS, 1):
        c = ws.cell(row=1, column=col, value=title)
        c.font = header_font
        c.fill = header_fill
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.freeze_panes = 'A2'

    bold = Font(bold=True)
    for i, t in enumerate(plan.tasks, 2):
        if t.is_summary and t.is_milestone:
            typ = 'Сводка (веха)'
        elif t.is_summary:
            typ = 'Сводка'
        elif t.is_milestone:
            typ = 'Веха'
        else:
            typ = 'Задача'
        row = [t.uid, t.wbs, t.name, typ,
               t.start, t.finish, t.duration_days,
               t.percent_complete or None,
               ';'.join(t.predecessors) or None,
               t.deadline, t.cost,
               t.baseline_start, t.baseline_finish,
               t.responsible or None]
        for col, value in enumerate(row, 1):
            c = ws.cell(row=i, column=col, value=value)
            if t.is_summary:
                c.font = bold
            if col in DATE_COLS and value is not None:
                c.number_format = 'DD.MM.YYYY'

    wb.save(path)
    return path
