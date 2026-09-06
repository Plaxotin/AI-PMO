#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Парсер плана (BL-1): .xlsx/.xls/.csv → Plan; .mpp → MPXJ-конвертер → Plan.

Маппинг колонок — по ИМЕНАМ заголовков (подход BL-6, COLUMN_SYNONYMS).
Колонки не на своих местах / переименованные варианты — не страшно.
"""

import csv
import os
import re
from datetime import date, datetime
from typing import Optional

from plan_model import Plan, Task

# Синонимы заголовков → каноническое поле Task.
COLUMN_SYNONYMS = {
    'uid': ['id', 'uid', 'ид', 'номер', '№'],
    'name': ['название', 'name', 'task name', 'имя задачи', 'задача', 'наименование'],
    'wbs': ['wbs', 'сдр', 'структурная декомпозиция'],
    'start': ['начало', 'start', 'дата начала', 'start date', 'план. начало'],
    'finish': ['окончание', 'finish', 'дата окончания', 'finish date', 'срок', 'план. окончание'],
    'duration': ['длительность', 'duration', 'длит.', 'дней'],
    'percent': ['% завершения', '% complete', 'процент завершения', '%', 'выполнение'],
    'predecessors': ['предшественники', 'predecessors', 'предш.', 'связи'],
    'deadline': ['крайний срок', 'deadline', 'ограничение'],
    'cost': ['затраты', 'cost', 'вес', 'стоимость'],
    'baseline_start': ['базовое начало', 'baseline start', 'баз. начало'],
    'baseline_finish': ['базовое окончание', 'baseline finish', 'баз. окончание'],
    'responsible': ['ответственный', 'ресурсы', 'resources', 'resource names', 'исполнитель'],
}

DATE_FORMATS = ('%d.%m.%Y', '%Y-%m-%d', '%d.%m.%y', '%m/%d/%Y', '%d/%m/%Y')


def _parse_date(value) -> Optional[date]:
    if value is None or value == '':
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    s = str(value).strip()
    for fmt in DATE_FORMATS:
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _parse_float(value) -> Optional[float]:
    if value is None or value == '':
        return None
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip().replace('%', '').replace(',', '.')
    m = re.search(r'-?\d+(\.\d+)?', s)
    return float(m.group(0)) if m else None


def _parse_pred_uids(value) -> list:
    """'3;5FS+2d;7' → ['3','5','7'] (типы связей и лаги отбрасываем)."""
    if not value:
        return []
    uids = []
    for part in re.split(r'[;,]', str(value)):
        m = re.match(r'\s*(\d+)', part)
        if m:
            uids.append(m.group(1))
    return uids


def _map_headers(headers: list) -> dict:
    """Имена заголовков → индекс колонки, по COLUMN_SYNONYMS."""
    mapping = {}
    lowered = [(str(h).strip().lower() if h else '') for h in headers]
    for canon, synonyms in COLUMN_SYNONYMS.items():
        for i, h in enumerate(lowered):
            if h and h in [s.lower() for s in synonyms]:
                mapping[canon] = i
                break
    return mapping


def _rows_to_plan(rows: list, headers: list, name: str, fmt: str) -> Plan:
    mapping = _map_headers(headers)
    if 'name' not in mapping:
        raise ValueError('Не найдена колонка с названием задачи (проверьте заголовки файла)')

    def cell(row, key):
        i = mapping.get(key)
        return row[i] if i is not None and i < len(row) else None

    tasks = []
    for idx, row in enumerate(rows, start=2):  # +1 строка заголовка
        task_name = cell(row, 'name')
        if task_name is None or str(task_name).strip() == '':
            continue
        uid = str(cell(row, 'uid') or idx - 1).strip()
        outline = 1
        # Уровень вложенности: по WBS (1.2.3 → 3) или по отступу в названии
        wbs = str(cell(row, 'wbs') or '').strip()
        if wbs:
            outline = wbs.count('.') + 1
        else:
            raw = str(task_name)
            stripped = raw.lstrip()
            outline = max(1, (len(raw) - len(stripped)) // 2 + 1)
        duration = _parse_float(cell(row, 'duration'))
        tasks.append(Task(
            uid=uid,
            name=str(task_name).strip(),
            wbs=wbs,
            outline_level=outline,
            start=_parse_date(cell(row, 'start')),
            finish=_parse_date(cell(row, 'finish')),
            duration_days=duration,
            percent_complete=_parse_float(cell(row, 'percent')) or 0.0,
            predecessors=_parse_pred_uids(cell(row, 'predecessors')),
            deadline=_parse_date(cell(row, 'deadline')),
            cost=_parse_float(cell(row, 'cost')),
            baseline_start=_parse_date(cell(row, 'baseline_start')),
            baseline_finish=_parse_date(cell(row, 'baseline_finish')),
            responsible=str(cell(row, 'responsible') or '').strip(),
            row_ref=f'строка {idx}',
        ))

    # summary = задача, за которой следует задача с бо́льшим outline_level;
    # а также строка-заголовок без длительности, дат и процента (типичный
    # этап в Excel-выгрузке: «Этап 1. …»)
    for i, t in enumerate(tasks):
        if i + 1 < len(tasks) and tasks[i + 1].outline_level > t.outline_level:
            t.is_summary = True
        elif (t.duration_days is None and not t.start and not t.finish
              and not t.percent_complete):
            t.is_summary = True
        elif duration_is_zero_milestone(t):
            t.is_milestone = True

    # successors — обратные связи из predecessors
    by_uid = {t.uid: t for t in tasks}
    for t in tasks:
        for p in t.predecessors:
            if p in by_uid:
                by_uid[p].successors.append(t.uid)

    return Plan(name=name, source_format=fmt, tasks=tasks,
                columns_found=sorted(mapping.keys()))


def duration_is_zero_milestone(t: Task) -> bool:
    return (t.duration_days == 0) and not t.is_summary


def parse_xlsx(path: str, name: str) -> Plan:
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        raise ValueError('Пустой лист Excel')
    headers, data = rows[0], rows[1:]
    return _rows_to_plan(data, list(headers), name, 'xlsx')


def parse_csv(path: str, name: str) -> Plan:
    with open(path, 'r', encoding='utf-8-sig', errors='replace') as f:
        sample = f.read(4096)
        f.seek(0)
        dialect = csv.Sniffer().sniff(sample, delimiters=';,\t') if sample.strip() else csv.excel
        rows = list(csv.reader(f, dialect))
    if not rows:
        raise ValueError('Пустой CSV')
    headers, data = rows[0], rows[1:]
    return _rows_to_plan(data, headers, name, 'csv')


def parse_mpp(path: str, name: str) -> Plan:
    """MPP → Plan через MPXJ (JVM поднимается один раз на процесс).

    Требует на сервере: default-jre-headless + pip-пакеты jpype1 и mpxj.
    """
    _ensure_jvm()
    import jpype
    reader = jpype.JClass('org.mpxj.reader.UniversalProjectReader')()
    return _mpxj_to_plan(reader.read(path), name)


_jvm_ready = False


def _ensure_jvm():
    global _jvm_ready
    if _jvm_ready:
        return
    if 'JAVA_HOME' not in os.environ and \
            os.path.exists('/usr/lib/jvm/default-java'):
        os.environ['JAVA_HOME'] = '/usr/lib/jvm/default-java'
    import jpype
    import mpxj  # noqa: F401 — регистрирует jar'ы MPXJ в classpath
    if not jpype.isJVMStarted():
        jpype.startJVM()
    _jvm_ready = True


def _ld(value) -> Optional[date]:
    """java.time.LocalDateTime/LocalDate → date (или None)."""
    if value is None:
        return None
    try:
        return date.fromisoformat(str(value)[:10])
    except ValueError:
        return None


def _dur_days(dur) -> Optional[float]:
    """MPXJ Duration → дни (8-часовой день, 5-дневная неделя)."""
    if dur is None:
        return None
    v = float(dur.getDuration())
    units = dur.getUnits()
    try:
        u = str(units.name())  # каноническое имя enum: HOURS, DAYS…
    except Exception:
        u = str(units).upper()
    table = {
        'MINUTES': v / 480.0, 'HOURS': v / 8.0, 'DAYS': v,
        'WEEKS': v * 5.0, 'MONTHS': v * 20.0, 'YEARS': v * 240.0,
        'ELAPSED_MINUTES': v / 1440.0, 'ELAPSED_HOURS': v / 24.0,
        'ELAPSED_DAYS': v, 'ELAPSED_WEEKS': v * 7.0,
        'ELAPSED_MONTHS': v * 30.0, 'ELAPSED_YEARS': v * 365.0,
    }
    return table.get(u, v)


_REL_TYPES = {'FINISH_START': 'FS', 'START_START': 'SS',
              'FINISH_FINISH': 'FF', 'START_FINISH': 'SF'}


def _mpxj_to_plan(pf, name: str) -> Plan:
    """Конвертирует org.mpxj.ProjectFile в Plan."""
    tasks = []
    for jt in pf.getTasks():
        if jt is None or bool(jt.getNull()):
            continue
        uid_v = jt.getUniqueID()
        if uid_v is None or int(uid_v) == 0:
            continue  # проектная сводка (UID 0)
        uid = str(int(uid_v))

        preds = []
        pred_detail = []
        try:
            for rel in jt.getPredecessors():
                tgt = rel.getPredecessorTask()
                if tgt is None or tgt.getUniqueID() is None:
                    continue
                puid = str(int(tgt.getUniqueID()))
                preds.append(puid)
                rtype = rel.getType()
                pred_detail.append({
                    'uid': puid,
                    'type': _REL_TYPES.get(str(rtype.name())
                                           if rtype is not None else '', 'FS'),
                    'lag_days': _dur_days(rel.getLag()) or 0.0,
                })
        except Exception:
            pass

        constraint = ''
        try:
            ct = jt.getConstraintType()
            if ct is not None:
                constraint = str(ct.name())
        except Exception:
            pass

        responsible = ''
        try:
            names = []
            for a in jt.getResourceAssignments():
                r = a.getResource()
                if r is not None and r.getName():
                    names.append(str(r.getName()))
            responsible = ', '.join(names[:3])
        except Exception:
            pass

        pct = jt.getPercentageComplete()
        cost = jt.getCost()
        level = jt.getOutlineLevel()
        tasks.append(Task(
            uid=uid,
            name=str(jt.getName() or ''),
            wbs=str(jt.getWBS() or ''),
            outline_level=int(level) if level is not None else 1,
            is_summary=bool(jt.getSummary()),
            is_milestone=bool(jt.getMilestone()),
            start=_ld(jt.getStart()),
            finish=_ld(jt.getFinish()),
            duration_days=_dur_days(jt.getDuration()),
            percent_complete=float(pct) if pct is not None else 0.0,
            predecessors=preds,
            pred_detail=pred_detail,
            constraint_type=constraint,
            deadline=_ld(jt.getDeadline()),
            cost=float(cost) if cost is not None else None,
            baseline_start=_ld(jt.getBaselineStart()),
            baseline_finish=_ld(jt.getBaselineFinish()),
            responsible=responsible,
            row_ref=f'mpp uid={uid}',
        ))

    by_uid = {t.uid: t for t in tasks}
    for t in tasks:
        for p in t.predecessors:
            if p in by_uid:
                by_uid[p].successors.append(t.uid)

    if not tasks:
        raise ValueError('В .mpp не найдено ни одной задачи')
    # columns_found — канонические имена полей, которые .mpp несёт всегда
    # (на них опираются проверки R-07/R-08: для .mpp состав колонок присущ формату)
    return Plan(name=name, source_format='mpp', tasks=tasks,
                columns_found=['uid', 'name', 'wbs', 'start', 'finish',
                               'duration', 'percent_complete', 'predecessors',
                               'deadline', 'cost', 'baseline_start',
                               'baseline_finish', 'responsible'])


def parse_plan(path: str, filename: Optional[str] = None) -> Plan:
    """Точка входа: определяет формат по расширению и возвращает Plan."""
    name = filename or os.path.basename(path)
    ext = os.path.splitext(name)[1].lower()
    if ext in ('.xlsx', '.xls'):
        return parse_xlsx(path, name)
    if ext == '.csv':
        return parse_csv(path, name)
    if ext == '.mpp':
        return parse_mpp(path, name)
    raise ValueError(f'Неподдерживаемый формат {ext}. Пришлите .xlsx, .csv или .mpp')
