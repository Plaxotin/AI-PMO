"""Формирование RACI-матрицы в Excel (openpyxl).

Один лист «RACI»: заголовок, метаданные проекта, матрица
(активности x роли), легенда R/A/C/I и примечания.
"""
from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ACCENT = "1F4E79"        # тёмно-синяя шапка
ACCENT_LIGHT = "DDEBF7"  # заливка ячеек A
PHASE_FILL = "F2F2F2"    # разделители фаз
WHITE = "FFFFFF"

THIN = Side(style="thin", color="B0B0B0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

LEGEND = [
    ("R", "Responsible — Исполнитель: выполняет работу (может быть несколько)"),
    ("A", "Accountable — Ответственный: принимает решение и отвечает за результат (строго один)"),
    ("C", "Consulted — Консультируемый: экспертиза и двусторонняя коммуникация до выполнения"),
    ("I", "Informed — Информируемый: одностороннее уведомление о ходе/результате"),
]


def build_xlsx(data: dict, path: str) -> str:
    wb = Workbook()
    ws = wb.active
    ws.title = "RACI"
    ws.sheet_view.showGridLines = False

    roles = data.get("roles") or []
    activities = data.get("activities") or []
    n_cols = len(roles) + 1  # + колонка активности

    # --- Заголовок (строка 2) ---
    ws.cell(row=2, column=2, value=f"RACI-матрица — {data.get('project_name', 'проект')}")
    ws.cell(row=2, column=2).font = Font(size=16, bold=True, color=ACCENT)
    ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=1 + n_cols)

    # --- Метаданные проекта ---
    meta = [
        ("Название проекта", data.get("project_name", "")),
        ("Цель проекта", data.get("project_goal", "")),
        ("Содержание / рамки", data.get("project_scope", "")),
        ("Сформировано", "AI PMO · BL-24 Генератор RACI (методология PMBOK)"),
    ]
    row = 4
    for label, value in meta:
        c1 = ws.cell(row=row, column=2, value=label)
        c1.font = Font(bold=True, color=ACCENT)
        c1.alignment = Alignment(vertical="top")
        c2 = ws.cell(row=row, column=3, value=value)
        c2.alignment = Alignment(wrap_text=True, vertical="top")
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=1 + n_cols)
        row += 1

    # --- Шапка матрицы ---
    row += 1
    header_row = row
    h = ws.cell(row=header_row, column=2, value="Фаза / Активность")
    for i, role in enumerate(roles):
        ws.cell(row=header_row, column=3 + i, value=role)
    for col in range(2, 2 + n_cols):
        c = ws.cell(row=header_row, column=col)
        c.font = Font(bold=True, color=WHITE, size=10)
        c.fill = PatternFill("solid", fgColor=ACCENT)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER
    ws.row_dimensions[header_row].height = 60

    # --- Тело матрицы ---
    prev_phase = None
    for act in activities:
        row += 1
        phase = act.get("phase", "")
        name = act.get("name", "")
        label = f"{phase} · {name}" if phase != prev_phase else name
        prev_phase = phase
        c = ws.cell(row=row, column=2, value=label)
        c.alignment = Alignment(wrap_text=True, vertical="center")
        c.font = Font(size=10, bold=(phase != prev_phase or " · " in label))
        if " · " in label:
            c.fill = PatternFill("solid", fgColor=PHASE_FILL)
        c.border = BORDER
        assign = act.get("assignments") or {}
        for i, role in enumerate(roles):
            v = assign.get(role, "")
            cell = ws.cell(row=row, column=3 + i, value=v)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.font = Font(size=10, bold=(v == "A"), color=ACCENT if v == "A" else "000000")
            cell.border = BORDER
            if v == "A":
                cell.fill = PatternFill("solid", fgColor=ACCENT_LIGHT)
            elif " · " in label:
                cell.fill = PatternFill("solid", fgColor=PHASE_FILL)
        ws.row_dimensions[row].height = 28

    # --- Легенда ---
    row += 2
    ws.cell(row=row, column=2, value="Легенда RACI (PMBOK)").font = Font(bold=True, color=ACCENT)
    for letter, text in LEGEND:
        row += 1
        c = ws.cell(row=row, column=2, value=letter)
        c.font = Font(bold=True, size=10)
        c.alignment = Alignment(horizontal="center")
        c.border = BORDER
        c.fill = PatternFill("solid", fgColor=ACCENT_LIGHT if letter == "A" else PHASE_FILL)
        t = ws.cell(row=row, column=3, value=text)
        t.alignment = Alignment(wrap_text=True, vertical="center")
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=1 + n_cols)

    # --- Примечания ---
    notes = data.get("notes") or []
    if notes:
        row += 2
        ws.cell(row=row, column=2, value="Примечания и допущения").font = Font(bold=True, color=ACCENT)
        for i, note in enumerate(notes, 1):
            row += 1
            c = ws.cell(row=row, column=2, value=f"{i}. {note}")
            c.alignment = Alignment(wrap_text=True, vertical="top")
            c.font = Font(size=9)
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=1 + n_cols)
            ws.row_dimensions[row].height = 24

    # --- Ширины и закрепление ---
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 42
    for i in range(len(roles)):
        ws.column_dimensions[get_column_letter(3 + i)].width = 14
    ws.freeze_panes = ws.cell(row=header_row + 1, column=3)

    wb.save(path)
    return path
